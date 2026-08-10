"""Collated BOM exporter plugin for InvenTree.

Adds a "Collated BOM Exporter" option to the standard BOM export dialog. Where
the built-in exporter lists the same part once per sub-assembly it appears in,
this one rolls every occurrence of a part into a single line with a true total
quantity for building one of the top-level assembly. Quantities are multiplied
down through the BOM tree, so a screw used four times inside a sub-assembly that
is itself used twice comes out as eight.

If stock data is included, each line also shows current stock, the shortfall and
whether there is enough on hand. If pricing is included, each line shows the
supplier, a unit price and a line total, with the order quantity rounded up to
the supplier pack / minimum order size (need 5, sold in packs of 10 -> order
10). A grand total sits at the bottom.

Requires InvenTree 0.17.0 or newer (the data export plugin framework).
"""

from collections import OrderedDict
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

import rest_framework.serializers as serializers

from InvenTree.helpers import normalize
from part.models import BomItem, Part
from part.serializers import BomItemSerializer
from plugin import InvenTreePlugin
from plugin.mixins import DataExportMixin, UrlsMixin


ZERO = Decimal(0)
ONE = Decimal(1)


def _dec(value, default=ZERO):
    """Coerce anything to a Decimal without ever raising."""
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default


def _round_up_to_pack(quantity, pack_size):
    """Round quantity up to the next whole multiple of pack_size.

    A pack_size of 10 with a need of 5 gives 10; a need of 11 gives 20. A
    pack_size of 1 (or missing) rounds up to the next whole unit.
    """
    quantity = _dec(quantity)
    pack_size = _dec(pack_size, ONE)

    if quantity <= ZERO:
        return ZERO
    if pack_size <= ZERO:
        pack_size = ONE

    packs = (quantity / pack_size).to_integral_value(rounding="ROUND_CEILING")
    return packs * pack_size


class CollatedBomOptionsSerializer(serializers.Serializer):
    """User-facing options shown in the BOM export dialog."""

    components_only = serializers.BooleanField(
        default=True,
        label=_('Components only'),
        help_text=_(
            'Exclude sub-assembly lines, keep only the parts you pull from the '
            'shelf. Sub-assembly quantities are still rolled into their parts.'
        ),
    )

    export_levels = serializers.IntegerField(
        default=0,
        label=_('Levels'),
        help_text=_('How many levels deep to explode. Zero means all levels.'),
        min_value=0,
    )

    export_stock_data = serializers.BooleanField(
        default=True,
        label=_('Stock data'),
        help_text=_('Include current stock, shortfall and a buildable flag.'),
    )

    export_pricing = serializers.BooleanField(
        default=True,
        label=_('Pricing'),
        help_text=_(
            'Include supplier, unit price, order quantity and line total, plus '
            'a grand total row.'
        ),
    )

    round_to_packs = serializers.BooleanField(
        default=True,
        label=_('Round up to pack size'),
        help_text=_(
            'Round the order quantity up to the supplier pack / minimum order '
            'size where it is recorded against the part.'
        ),
    )


class CollatedBomExporterPlugin(DataExportMixin, UrlsMixin, InvenTreePlugin):
    """Export a BOM with identical parts collated into one line each."""

    NAME = 'Collated BOM Exporter'
    SLUG = 'collated-bom-exporter'
    TITLE = _('Collated BOM Exporter')
    DESCRIPTION = _(
        'Exports a BOM with duplicate parts collated across sub-assemblies into '
        'a single line each, with true total quantities, a stock check and '
        'pack-rounded order pricing.'
    )
    VERSION = '0.4.0'
    AUTHOR = _('Contour3D')

    ExportOptionsSerializer = CollatedBomOptionsSerializer

    def supports_export(self, model_class: type, user, *args, **kwargs) -> bool:
        """Offer this exporter only for BOM data."""
        return (
            model_class == BomItem
            and kwargs.get('serializer_class') == BomItemSerializer
        )

    def generate_filename(self, model_class, export_format: str) -> str:
        """Name the output file."""
        date = timezone.now().date().isoformat()
        return f'InvenTree_Collated_BOM_{date}.{export_format}'

    def prefetch_queryset(self, queryset):
        """Reduce query count when walking the tree."""
        return queryset.prefetch_related('sub_part')

    def export_data(
        self, queryset, serializer_class, headers, context, output, **kwargs
    ):
        """Walk the BOM tree, collate by part, and return one row per part."""
        self.serializer_class = serializer_class

        # Cache the chosen options for use here and in update_headers.
        self.components_only = context.get('components_only', True)
        self.export_levels = context.get('export_levels', 0)
        self.export_stock_data = context.get('export_stock_data', True)
        self.export_pricing = context.get('export_pricing', True)
        self.round_to_packs = context.get('round_to_packs', True)

        rows, grand_total = self._collate_rows(self.prefetch_queryset(queryset))

        # Grand total row at the bottom.
        if self.export_pricing:
            total_row = {key: '' for key in headers.keys()}
            total_row['part'] = str(_('TOTAL'))
            total_row['line_total'] = normalize(grand_total)
            rows.append(total_row)

        return rows

    def _collate_rows(self, top_items):
        """Collate the BOM tree under the given top-level items.

        Returns (rows, grand_total) where rows is the finalised list of dicts
        (without the TOTAL row). Shared by the export dialog and the styled
        download so both produce identical numbers.
        """
        # part id -> collated row
        self.collated: "OrderedDict[int, dict]" = OrderedDict()

        for bom_item in top_items:
            self._process(bom_item, level=1, multiplier=ONE)

        rows = list(self.collated.values())
        grand_total = ZERO

        # Finalise derived fields now that totals are complete.
        for row in rows:
            total = _dec(row['_total'])
            row['total_quantity'] = normalize(total)
            row['bom_level'] = row['_level']

            available = row.get('_available')
            short = None
            if available is not None:
                short = total - _dec(available)
                if short < ZERO:
                    short = ZERO

            if self.export_stock_data:
                if available is None:
                    row['available_stock'] = ''
                    row['shortfall'] = ''
                    row['buildable'] = ''
                else:
                    row['available_stock'] = normalize(_dec(available))
                    row['shortfall'] = normalize(short) if short > ZERO else 0
                    row['buildable'] = _('No') if short > ZERO else _('Yes')

            if self.export_pricing:
                # Order the shortfall where we know stock, otherwise the whole
                # requirement. Round up to the pack size if asked.
                need = short if short is not None else total
                pack = _dec(row.get('_pack'), ONE) if self.round_to_packs else ONE
                order_qty = (
                    _round_up_to_pack(need, pack)
                    if self.round_to_packs
                    else need
                )

                unit_price = row.get('_unit_price')
                row['supplier'] = row.get('_supplier', '')
                row['supplier_sku'] = row.get('_sku', '')
                row['pack_size'] = normalize(pack)
                row['order_quantity'] = normalize(order_qty)

                lead = row.get('_lead_time')
                row['lead_time'] = normalize(_dec(lead)) if lead is not None else ''

                if unit_price is not None:
                    line_total = _dec(unit_price) * _dec(order_qty)
                    grand_total += line_total
                    row['unit_price'] = normalize(_dec(unit_price))
                    row['line_total'] = normalize(line_total)
                else:
                    row['unit_price'] = ''
                    row['line_total'] = ''

            row['reference'] = ', '.join(sorted(r for r in row['_refs'] if r))

        # Buyable-first when we know stock, then by name; otherwise by name.
        if self.export_stock_data:
            rows.sort(
                key=lambda r: (
                    _dec(r.get('shortfall') or 0) <= ZERO,
                    str(r.get('part', '')).lower(),
                )
            )
        else:
            rows.sort(key=lambda r: str(r.get('part', '')).lower())

        # Strip the private working fields before returning.
        for row in rows:
            for key in ('_total', '_available', '_refs', '_level', '_pack',
                        '_unit_price', '_supplier', '_sku', '_lead_time'):
                row.pop(key, None)

        return rows, grand_total

    def _process(
        self,
        bom_item: BomItem,
        level: int,
        multiplier: Decimal,
        **kwargs,
    ) -> None:
        """Recursively process one BOM item and its children."""
        part = bom_item.sub_part
        total = _dec(bom_item.quantity) * multiplier
        is_assembly = bool(part.assembly)

        include = not (self.components_only and is_assembly)
        if include:
            self._accumulate(bom_item, part, total, level)

        # Recurse into sub-assemblies unless we have hit the level limit.
        if is_assembly and (self.export_levels <= 0 or level < self.export_levels):
            sub_items = part.get_bom_items()
            sub_items = self.prefetch_queryset(sub_items)
            sub_items = BomItemSerializer.annotate_queryset(sub_items)
            for item in sub_items.all():
                self._process(
                    item,
                    level=level + 1,
                    multiplier=multiplier * _dec(bom_item.quantity),
                    **kwargs,
                )

    def _accumulate(self, bom_item: BomItem, part, total: Decimal, level: int) -> None:
        """Add this occurrence into the collated row for its part."""
        key = part.pk
        existing = self.collated.get(key)

        if existing:
            existing['_total'] += total
            existing['occurrences'] += 1
            existing['_level'] = min(existing['_level'], level)
            if bom_item.reference:
                existing['_refs'].add(bom_item.reference)
            return

        row: dict = {
            'part': part.name,
            'ipn': part.IPN or '',
            'description': part.description or '',
            'occurrences': 1,
            '_total': total,
            '_level': level,
            '_refs': {bom_item.reference} if bom_item.reference else set(),
        }

        # Pull stock and pricing from the BOM serializer once (InvenTree
        # annotates available_stock and pricing onto the queryset).
        if self.export_stock_data or self.export_pricing:
            data = self.serializer_class(bom_item, exporting=True).data

            if self.export_stock_data:
                available = data.get('available_stock')
                row['_available'] = available if available is not None else None

            if self.export_pricing:
                row['_unit_price'] = self._unit_price(data)
                info = self._supplier_info(part)
                row['_supplier'] = info['supplier']
                row['_sku'] = info['sku']
                row['_pack'] = info['pack']
                # Lead time lives on the part (how long the item takes), so read
                # it from the part regardless of whether a supplier is set.
                row['_lead_time'] = self._lead_time(part, info['supplier_part'])

        self.collated[key] = row

    @staticmethod
    def _unit_price(serializer_data) -> "Decimal | None":
        """Per-unit price from InvenTree's computed BOM pricing.

        Uses the minimum price where available, falling back to the maximum.
        """
        for field in ('pricing_min', 'pricing_max'):
            value = serializer_data.get(field)
            if value is not None:
                price = _dec(value, default=None) if value is not None else None
                if price is not None and price > ZERO:
                    return price
        return None

    @staticmethod
    def _supplier_info(part):
        """Cheapest supplier's name, SKU and pack size for a part.

        Returns sensible blanks when no supplier is recorded, so the export
        never fails on an incompletely set up part. 'supplier_part' is the
        chosen SupplierPart (or None) for the lead-time lookup.
        """
        blank = {'supplier': '', 'sku': '', 'pack': ONE, 'supplier_part': None}

        try:
            supplier_parts = list(part.supplier_parts.all())
        except Exception:
            supplier_parts = []

        if not supplier_parts:
            return blank

        best = None
        best_price = None
        for sp in supplier_parts:
            price = CollatedBomExporterPlugin._cheapest_price(sp)
            if price is None:
                continue
            if best_price is None or price < best_price:
                best_price = price
                best = sp

        # Fall back to the first supplier so the buyer still has a lead.
        if best is None:
            best = supplier_parts[0]

        supplier = getattr(best, 'supplier', None)
        return {
            'supplier': getattr(supplier, 'name', '') or '',
            'sku': getattr(best, 'SKU', '') or '',
            'pack': CollatedBomExporterPlugin._pack_size(best),
            'supplier_part': best,
        }

    @staticmethod
    def _lead_time(part, supplier_part=None):
        """Lead time in days for a part, from wherever it is recorded.

        Lead time is treated as a property of the PART (how long that item
        takes to get), so the part is checked first and works even when the
        part has no supplier. Checks, in order:
          1. a Part parameter named "Lead Time" / "Lead Time (days)",
          2. Part.metadata (key 'lead_time' / 'lead_time_days'),
          3. a "lead_time: 14" tag in the part's notes,
          4. the native SupplierPart.lead_time attribute (days or a duration),
          5. SupplierPart.metadata,
          6. a "lead_time: 14" tag in the supplier part note.

        Returns a whole number of days, or None if nothing is recorded.
        """
        # 1. Part parameter (where Charlotte records it).
        value = CollatedBomExporterPlugin._part_parameter_days(part)
        if value is not None:
            return value

        # 2. Part metadata.
        value = CollatedBomExporterPlugin._metadata_days(part)
        if value is not None:
            return value

        # 3. A tag in the part notes.
        note = getattr(part, 'notes', None) or getattr(part, 'note', None) or ''
        value = CollatedBomExporterPlugin._parse_lead_tag(str(note))
        if value is not None:
            return value

        if supplier_part is None:
            return None

        # 4. Native SupplierPart attribute (exists but usually no UI).
        raw = getattr(supplier_part, 'lead_time', None)
        days = getattr(raw, 'days', None)  # timedelta support
        value = (
            _dec(days, default=None) if days is not None
            else _dec(raw, default=None) if raw not in (None, '')
            else None
        )
        if value is not None and value > ZERO:
            return value

        # 5. Supplier part metadata.
        value = CollatedBomExporterPlugin._metadata_days(supplier_part)
        if value is not None:
            return value

        # 6. A tag in the supplier part note.
        note = getattr(supplier_part, 'note', None) or getattr(supplier_part, 'notes', None) or ''
        return CollatedBomExporterPlugin._parse_lead_tag(str(note))

    @staticmethod
    def _metadata_days(obj):
        """Read a lead-time value from a model's metadata dict, if present."""
        try:
            meta = obj.metadata or {}
        except Exception:
            return None
        for key in ('lead_time', 'lead_time_days', 'leadTime'):
            if key in meta:
                value = _dec(meta[key], default=None)
                if value is not None and value > ZERO:
                    return value
        return None

    @staticmethod
    def _parse_lead_tag(text):
        """Pull a lead-time number out of free text like 'lead_time: 14'."""
        import re

        match = re.search(
            r'lead[\s_]*(?:time)?\s*[:=]?\s*(\d+(?:\.\d+)?)',
            text,
            flags=re.IGNORECASE,
        )
        if match:
            value = _dec(match.group(1), default=None)
            if value is not None and value > ZERO:
                return value
        return None

    @staticmethod
    def _part_parameter_days(part):
        """Read a 'Lead Time' part parameter, if one is set.

        Matches any parameter whose name contains both 'lead' and 'time', and
        copes with values stored as a plain number, a numeric field, or text
        like '21 days'.
        """
        try:
            params = part.get_parameters()
        except Exception:
            try:
                params = part.parameters.all()
            except Exception:
                params = []

        for param in params:
            template = getattr(param, 'template', None)
            name = str(getattr(template, 'name', '') or '').lower()
            if 'lead' in name and 'time' in name:
                # Prefer InvenTree's parsed numeric value when available.
                value = _dec(getattr(param, 'data_numeric', None), default=None)
                if value is None:
                    value = CollatedBomExporterPlugin._first_number(
                        getattr(param, 'data', None)
                    )
                if value is not None and value > ZERO:
                    return value
        return None

    @staticmethod
    def _first_number(value):
        """Pull the first number out of a value like '21' or '21 days'."""
        if value is None:
            return None
        import re

        match = re.search(r'\d+(?:\.\d+)?', str(value))
        return _dec(match.group(0), default=None) if match else None

    @staticmethod
    def _cheapest_price(supplier_part):
        """Lowest per-unit price across a supplier part's price breaks."""
        breaks = []
        for attr in ('pricebreaks', 'price_breaks'):
            manager = getattr(supplier_part, attr, None)
            if manager is not None:
                try:
                    breaks = list(manager.all())
                    break
                except Exception:
                    breaks = []

        pack = CollatedBomExporterPlugin._pack_size(supplier_part)
        pack = pack if pack > ZERO else ONE

        cheapest = None
        for pb in breaks:
            raw = getattr(pb, 'price', None)
            if raw is None:
                continue
            price = _dec(raw, default=None)
            if price is None:
                continue
            per_unit = price / pack
            if cheapest is None or per_unit < cheapest:
                cheapest = per_unit
        return cheapest

    @staticmethod
    def _pack_size(supplier_part):
        """Pack / minimum order multiple for a supplier part (defaults to 1)."""
        for attr in ('pack_quantity_native', 'multiple'):
            raw = getattr(supplier_part, attr, None)
            if raw is None:
                continue
            value = _dec(raw, default=None)
            if value is not None and value > ZERO:
                return value
        return ONE

    def update_headers(self, headers, context, **kwargs):
        """Replace the default columns with the collated column set.

        IPN is dropped (not needed) and Description is placed on the far right
        so the quantities and pricing read first.
        """
        columns: "OrderedDict[str, str]" = OrderedDict()
        columns['part'] = _('Part')
        columns['bom_level'] = _('BOM Level')
        columns['total_quantity'] = _('Total Quantity Required')

        if self.export_stock_data:
            columns['available_stock'] = _('Current Stock')
            columns['shortfall'] = _('Shortfall')
            columns['buildable'] = _('Enough In Stock')

        if self.export_pricing:
            columns['supplier'] = _('Supplier')
            columns['supplier_sku'] = _('Supplier SKU')
            columns['pack_size'] = _('Pack Size')
            columns['order_quantity'] = _('Order Qty')
            columns['unit_price'] = _('Unit Price')
            columns['line_total'] = _('Line Total')
            columns['lead_time'] = _('Lead Time (days)')

        columns['occurrences'] = _('Recurrences')
        columns['description'] = _('Description')
        return columns

    # ---- One-click styled download -----------------------------------------
    # This is the "always looks like that" path. The export dialog can only
    # write plain data, so a fully styled workbook (colours, tick column,
    # est. delivery) is built here with openpyxl and served as a download.

    DOWNLOAD_COLUMNS = [
        ('ordered', _('Ordered')),
        ('part', _('Part')),
        ('bom_level', _('BOM Level')),
        ('total_quantity', _('Total Quantity Required')),
        ('available_stock', _('Current Stock')),
        ('shortfall', _('Shortfall')),
        ('buildable', _('Enough In Stock')),
        ('supplier', _('Supplier')),
        ('order_quantity', _('Order Qty')),
        ('unit_price', _('Unit Price')),
        ('line_total', _('Line Total')),
        ('lead_time', _('Lead Time (days)')),
        ('est_delivery', _('Est. Delivery (if ordered today)')),
        ('occurrences', _('Recurrences')),
        ('description', _('Description')),
    ]

    LEAD_HELP = (
        "Lead time is how long the item takes to get, in days. Record it on the "
        "Part (Part > Parameters: 'Lead Time (days)', or Part > Notes: "
        "'lead_time: 14'). Amber cells have none set yet."
    )

    def setup_urls(self):
        """Expose the styled download at /plugin/collated-bom-exporter/."""
        from django.urls import path

        return [
            path('download/<int:pk>/', self.download_formatted_bom,
                 name='collated-bom-download'),
        ]

    def download_formatted_bom(self, request, pk):
        """Build and return the fully styled collated BOM for a part."""
        try:
            part = Part.objects.get(pk=pk)
        except Part.DoesNotExist:
            return HttpResponse('Part not found', status=404)

        # One-click download includes everything.
        self.serializer_class = BomItemSerializer
        self.components_only = True
        self.export_levels = 0
        self.export_stock_data = True
        self.export_pricing = True
        self.round_to_packs = True

        items = BomItemSerializer.annotate_queryset(
            self.prefetch_queryset(part.get_bom_items())
        )
        rows, grand_total = self._collate_rows(items)

        content = self._build_workbook_bytes(rows, grand_total)
        date = timezone.now().date().isoformat()
        safe_name = ''.join(c if c.isalnum() or c in '-_' else '_' for c in str(part.name))
        filename = f'Collated_BOM_{safe_name}_{date}.xlsx'
        response = HttpResponse(
            content,
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _build_workbook_bytes(self, rows, grand_total):
        """Render the collated rows into a styled .xlsx and return the bytes."""
        import io

        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        keys = [k for k, _label in self.DOWNLOAD_COLUMNS]
        labels = [str(label) for _k, label in self.DOWNLOAD_COLUMNS]
        idx = {k: i for i, k in enumerate(keys)}
        tick = '✔'

        def num(v):
            if v is None or v == '':
                return None
            if isinstance(v, Decimal):
                return float(v)
            if isinstance(v, (int, float)):
                return v
            try:
                return float(str(v))
            except (ValueError, TypeError):
                return None

        wb = Workbook()
        ws = wb.active
        ws.title = 'Collated BOM'
        ws.append(labels)

        for row in rows:
            out = []
            for k in keys:
                if k in ('ordered', 'est_delivery'):
                    out.append('')
                    continue
                v = row.get(k, '')
                if isinstance(v, Decimal):
                    v = float(v)
                elif k == 'buildable' and v != '':
                    v = str(v)
                out.append(v)
            ws.append(out)

        # TOTAL row: money summed, lead time as the longest (critical path).
        total = [''] * len(keys)
        total[idx['part']] = 'TOTAL'
        total[idx['line_total']] = float(grand_total) if grand_total else 0
        leads = [num(r.get('lead_time')) for r in rows]
        leads = [x for x in leads if x is not None]
        if leads:
            total[idx['lead_time']] = max(leads)
        ws.append(total)

        n = len(keys)
        first, last = 2, 1 + len(rows)
        total_row = ws.max_row

        header_fill = PatternFill('solid', fgColor='1F2A37')
        header_font = Font(bold=True, color='FFFFFF')
        green_fill = PatternFill('solid', fgColor='C6EFCE')
        green_font = Font(color='006100')
        red_fill = PatternFill('solid', fgColor='FFC7CE')
        red_font = Font(color='9C0006')
        amber_fill = PatternFill('solid', fgColor='FFF2CC')
        amber_font = Font(color='9C6500', italic=True)
        ordered_fill = PatternFill('solid', fgColor='E8EAED')
        ordered_font = Font(strike=True, color='6B7280')
        top_border = Border(top=Side(style='thin', color='9AA0A6'))

        for c in range(1, n + 1):
            hc = ws.cell(row=1, column=c)
            hc.fill = header_fill
            hc.font = header_font
            hc.alignment = Alignment(vertical='center', wrap_text=True)
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = f'A1:{get_column_letter(n)}1'

        for key in ('unit_price', 'line_total'):
            c = idx[key] + 1
            for r in range(2, total_row + 1):
                ws.cell(row=r, column=c).number_format = '$#,##0.00'

        if rows:
            # Enough In Stock: solid fills (always visible), plus a live rule
            # so edits recolour too. Green = Yes, red = No.
            bcol = idx['buildable'] + 1
            for r in range(first, last + 1):
                cellv = ws.cell(row=r, column=bcol)
                token = str(cellv.value).strip().lower() if cellv.value not in (None, '') else ''
                if token == 'yes':
                    cellv.fill, cellv.font = green_fill, green_font
                elif token == 'no':
                    cellv.fill, cellv.font = red_fill, red_font
            col = get_column_letter(bcol)
            rng = f'{col}{first}:{col}{last}'
            ws.conditional_formatting.add(rng, CellIsRule(
                operator='equal', formula=['"Yes"'], fill=green_fill, font=green_font))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator='equal', formula=['"No"'], fill=red_fill, font=red_font))

            # Ordered tick column + grey-out-when-ticked rule.
            oc = get_column_letter(idx['ordered'] + 1)
            dv = DataValidation(type='list', formula1=f'"{tick}"', allow_blank=True)
            dv.promptTitle = 'Ordered?'
            dv.prompt = 'Pick the tick to mark this line as ordered'
            ws.add_data_validation(dv)
            dv.add(f'{oc}{first}:{oc}{last}')
            for r in range(first, last + 1):
                ws.cell(row=r, column=idx['ordered'] + 1).alignment = Alignment(
                    horizontal='center')
            ws.conditional_formatting.add(
                f'A{first}:{get_column_letter(n)}{last}',
                FormulaRule(formula=[f'${oc}{first}="{tick}"'],
                            fill=ordered_fill, font=ordered_font),
            )

            # Est. Delivery = today + lead time; amber where no lead time set.
            lc = get_column_letter(idx['lead_time'] + 1)
            dc_i = idx['est_delivery'] + 1
            ws.cell(row=1, column=idx['lead_time'] + 1).comment = Comment(
                self.LEAD_HELP, 'BOM Collator')
            for r in range(first, last + 1):
                lead_cell = ws.cell(row=r, column=idx['lead_time'] + 1)
                dc = ws.cell(row=r, column=dc_i)
                dc.value = (
                    f'=IF(${lc}{r}="","Add lead time to the Part",TODAY()+${lc}{r})'
                )
                dc.number_format = 'dd/mm/yyyy'
                dc.alignment = Alignment(horizontal='center')
                if num(lead_cell.value) is None:
                    lead_cell.fill = amber_fill
                    dc.fill = amber_fill
                    dc.font = amber_font

            total_lead = ws.cell(row=total_row, column=idx['lead_time'] + 1)
            if num(total_lead.value) is not None:
                td = ws.cell(row=total_row, column=dc_i)
                td.value = f'=TODAY()+${lc}{total_row}'
                td.number_format = 'dd/mm/yyyy'
                td.alignment = Alignment(horizontal='center')

        for c in range(1, n + 1):
            tc = ws.cell(row=total_row, column=c)
            tc.font = Font(bold=True)
            tc.border = top_border
            if keys[c - 1] in ('unit_price', 'line_total'):
                tc.number_format = '$#,##0.00'

        for c, key in enumerate(keys, start=1):
            if key == 'ordered':
                ws.column_dimensions[get_column_letter(c)].width = 9
                continue
            longest = len(labels[c - 1])
            for r in range(2, total_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and not str(v).startswith('='):
                    longest = max(longest, len(str(v)))
            cap = 45 if key == 'description' else 30
            ws.column_dimensions[get_column_letter(c)].width = min(
                max(longest + 2, 10), cap)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
