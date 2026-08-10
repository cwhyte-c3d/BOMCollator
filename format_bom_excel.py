"""Tidy up a Collated BOM Exporter spreadsheet for reading and ordering.

InvenTree writes plain data, so the nice-to-haves live here. Run this on an
exported .xlsx and it will:

  - add an "Ordered" tick column you click to mark a line as ordered (the
    whole row then greys out and strikes through),
  - drop the IPN column,
  - move Description to the far right so the numbers read first,
  - colour "Enough In Stock" green for Yes, red for No,
  - add "Est. Delivery" = today + Lead Time (days) so you can see, at a glance,
    when each line would land if ordered now,
  - format Unit Price and Line Total as currency,
  - rebuild a bold TOTAL row (total quantity, order qty and total cost),
  - bold the header, freeze it, add a filter, and size the columns.

A note on the tick column: Excel's own click-checkboxes cannot be written by a
script, so "Ordered" is a dropdown - click the cell, pick the tick. It behaves
like a checkbox for filtering and the row greys out when ticked.

Usage:
    python format_bom_excel.py "path/to/export.xlsx"
    python format_bom_excel.py "path/to/export.xlsx" "path/to/output.xlsx"
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# --- Column names -----------------------------------------------------------
DROP_COLUMNS = ["IPN"]
DESCRIPTION_COLUMN = "Description"
STOCK_FLAG_COLUMN = "Enough In Stock"
LEAD_COLUMN = "Lead Time (days)"
DELIVERY_COLUMN = "Est. Delivery (if ordered today)"
ORDERED_COLUMN = "Ordered"
CURRENCY_COLUMNS = ["Unit Price", "Line Total"]
SUM_COLUMNS = ["Total Quantity Required", "Order Qty", "Line Total"]
TOTAL_LABEL = "TOTAL"
TICK = "✔"  # heavy check mark

# --- Styling ----------------------------------------------------------------
CURRENCY_FORMAT = "$#,##0.00"
DATE_FORMAT = "dd/mm/yyyy"

HEADER_FILL = PatternFill("solid", fgColor="1F2A37")   # dark slate
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
ORDERED_FILL = PatternFill("solid", fgColor="E8EAED")   # light grey
ORDERED_FONT = Font(strike=True, color="6B7280")
MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")   # amber
MISSING_FONT = Font(color="9C6500", italic=True)
TOTAL_FONT = Font(bold=True)
TOP_BORDER = Border(top=Side(style="thin", color="9AA0A6"))

YES_VALUES = {"yes", "y", "true", "1"}
NO_VALUES = {"no", "n", "false", "0"}

MISSING_LEAD_TEXT = "Add lead time to the Part"
LEAD_HOWTO = (
    "Lead time is how long the item takes to get, in days.\n"
    "It is recorded on the PART (different items take different times), so add "
    "it one of these ways and the exporter will pick it up:\n"
    "  1. Part > Parameters: add 'Lead Time (days)' with the number of days\n"
    "  2. Part > Notes: type  lead_time: 14  (days)\n"
    "Not every part will have one. Rows highlighted amber have none set yet."
)


def _to_number(value):
    """Return a float for numeric-looking values, else None."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except (ValueError, AttributeError):
        return None


def format_workbook(input_path: Path, output_path: Path) -> None:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Spreadsheet is empty.")

    header = list(rows[0])
    data = [list(r) for r in rows[1:]]

    # Drop any pre-existing TOTAL row so we do not double count.
    if data and str(data[-1][0]).strip().upper() == TOTAL_LABEL:
        data.pop()

    # New column order: Ordered first, drop unwanted columns, Est. Delivery
    # after Lead Time, Description last. Lead Time and Est. Delivery are always
    # present, even if the export predates them, so the "add it" reminder shows.
    kept = [h for h in header if h not in DROP_COLUMNS]
    if LEAD_COLUMN not in kept:
        kept.append(LEAD_COLUMN)
    if DESCRIPTION_COLUMN in kept:
        kept.remove(DESCRIPTION_COLUMN)
        kept.append(DESCRIPTION_COLUMN)
    kept.insert(kept.index(LEAD_COLUMN) + 1, DELIVERY_COLUMN)
    kept.insert(0, ORDERED_COLUMN)

    src_index = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = src_index.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = openpyxl.Workbook()
    sheet = out.active
    sheet.title = "Collated BOM"
    sheet.append(kept)

    # Added columns (Ordered, Est. Delivery) have no source data.
    for row in data:
        sheet.append(["" if name in (ORDERED_COLUMN, DELIVERY_COLUMN)
                      else cell(row, name) for name in kept])

    # TOTAL row.
    totals = {}
    for name in SUM_COLUMNS:
        if name in kept:
            total = sum(
                n for n in (_to_number(cell(r, name)) for r in data) if n is not None
            )
            totals[name] = total

    total_row = [round(totals[name], 4) if name in totals else None for name in kept]
    # Put the TOTAL label in the Part column (or the first column as a fallback).
    label_col = kept.index("Part") if "Part" in kept else 0
    total_row[label_col] = TOTAL_LABEL
    sheet.append(total_row)

    missing = _apply_styles(sheet, kept, data_rows=len(data))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(output_path)
    return missing


def _apply_styles(sheet, header, data_rows) -> int:
    col_index = {name: i + 1 for i, name in enumerate(header)}
    last_row = sheet.max_row
    first_data, last_data = 2, 1 + data_rows       # data rows only
    total_row_idx = last_row

    # Header.
    for c in range(1, len(header) + 1):
        hc = sheet.cell(row=1, column=c)
        hc.fill = HEADER_FILL
        hc.font = HEADER_FONT
        hc.alignment = Alignment(vertical="center", wrap_text=True)

    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    # Currency columns.
    for name in CURRENCY_COLUMNS:
        c = col_index.get(name)
        if not c:
            continue
        for r in range(2, last_row + 1):
            sheet.cell(row=r, column=c).number_format = CURRENCY_FORMAT

    # Colour the stock flag on data rows.
    flag_c = col_index.get(STOCK_FLAG_COLUMN)
    if flag_c and data_rows:
        for r in range(first_data, last_data + 1):
            cellv = sheet.cell(row=r, column=flag_c)
            token = str(cellv.value).strip().lower() if cellv.value is not None else ""
            if token in YES_VALUES:
                cellv.fill, cellv.font = GREEN_FILL, GREEN_FONT
            elif token in NO_VALUES:
                cellv.fill, cellv.font = RED_FILL, RED_FONT

    # Ordered tick column: a dropdown per data row.
    ordered_c = col_index.get(ORDERED_COLUMN)
    if ordered_c and data_rows:
        col = get_column_letter(ordered_c)
        dv = DataValidation(type="list", formula1=f'"{TICK}"', allow_blank=True)
        dv.prompt = "Pick the tick to mark this line as ordered"
        dv.promptTitle = "Ordered?"
        sheet.add_data_validation(dv)
        dv.add(f"{col}{first_data}:{col}{last_data}")
        for r in range(first_data, last_data + 1):
            sheet.cell(row=r, column=ordered_c).alignment = Alignment(
                horizontal="center"
            )

        # When ticked, grey out and strike through the whole row.
        row_range = f"A{first_data}:{get_column_letter(len(header))}{last_data}"
        rule = FormulaRule(
            formula=[f'${col}{first_data}="{TICK}"'],
            fill=ORDERED_FILL,
            font=ORDERED_FONT,
        )
        sheet.conditional_formatting.add(row_range, rule)

    # Est. Delivery = today + lead time, as a live formula. Where lead time is
    # missing, the cell shows a reminder and the lead-time cell is flagged amber.
    lead_c = col_index.get(LEAD_COLUMN)
    delivery_c = col_index.get(DELIVERY_COLUMN)
    missing = 0
    if lead_c and delivery_c and data_rows:
        lead_col = get_column_letter(lead_c)

        # How-to note on both headers so it is easy to find.
        note = Comment(LEAD_HOWTO, "BOM Collator")
        note.width, note.height = 340, 150
        sheet.cell(row=1, column=lead_c).comment = note
        sheet.cell(row=1, column=delivery_c).comment = Comment(LEAD_HOWTO, "BOM Collator")

        for r in range(first_data, last_data + 1):
            lead_cell = sheet.cell(row=r, column=lead_c)
            dc = sheet.cell(row=r, column=delivery_c)
            dc.value = (
                f'=IF(${lead_col}{r}="","{MISSING_LEAD_TEXT}",TODAY()+${lead_col}{r})'
            )
            dc.number_format = DATE_FORMAT
            dc.alignment = Alignment(horizontal="center")

            has_lead = _to_number(lead_cell.value) is not None
            if not has_lead:
                missing += 1
                lead_cell.fill = MISSING_FILL
                dc.fill = MISSING_FILL
                dc.font = MISSING_FONT

    # TOTAL row.
    for c in range(1, len(header) + 1):
        tc = sheet.cell(row=total_row_idx, column=c)
        tc.font = TOTAL_FONT
        tc.border = TOP_BORDER
        if header[c - 1] in CURRENCY_COLUMNS:
            tc.number_format = CURRENCY_FORMAT

    # Column widths.
    for c, name in enumerate(header, start=1):
        if name == ORDERED_COLUMN:
            sheet.column_dimensions[get_column_letter(c)].width = 9
            continue
        longest = len(str(name))
        for r in range(2, last_row + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None and not str(val).startswith("="):
                longest = max(longest, len(str(val)))
        width = min(max(longest + 2, 10), 55)
        sheet.column_dimensions[get_column_letter(c)].width = width
        if name == DESCRIPTION_COLUMN:
            for r in range(2, last_row + 1):
                sheet.cell(row=r, column=c).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    return missing


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python format_bom_excel.py <input.xlsx> [output.xlsx]")

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        raise SystemExit(f"File not found: {input_path}")

    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_path = input_path.with_name(f"{input_path.stem} (formatted).xlsx")

    missing = format_workbook(input_path, output_path)
    print(f"Written: {output_path}")
    if missing:
        print(
            f"Reminder: {missing} part(s) have no lead time set. "
            "Add it in InvenTree (Supplier Part Notes: 'lead_time: 14', or a "
            "'Lead Time (days)' part parameter). The amber cells in the sheet "
            "show which ones."
        )


if __name__ == "__main__":
    main()
